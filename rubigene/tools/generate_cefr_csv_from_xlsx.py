import os
import sys
import csv
import openpyxl

XLSX_PATH = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "CEFR-J Wordlist Ver1.6.xlsx"))
OUT_PATH = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "data", "cefr.csv"))
COLUMNS = ["word", "CEFR_level", "pos"]

# 許容する列名の揺れ
WORD_KEYS = {"word", "lemma", "headword", "単語"}
CEFR_KEYS = {"cefr", "cefr level", "level", "cefr_level", "レベル"}
POS_KEYS = {"pos", "part of speech", "品詞"}

# ヘッダー行を自動検出し、列インデックスを返す
def detect_header_and_indices(ws):
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if not row:
            continue
        headers = [str(cell).strip().lower() if cell else '' for cell in row]
        word_idx = next((j for j, h in enumerate(headers) if h in WORD_KEYS), None)
        cefr_idx = next((j for j, h in enumerate(headers) if h in CEFR_KEYS), None)
        pos_idx = next((j for j, h in enumerate(headers) if h in POS_KEYS), None)
        if word_idx is not None and cefr_idx is not None and pos_idx is not None:
            return i, word_idx, cefr_idx, pos_idx
    return None, None, None, None

def main():
    if not os.path.exists(XLSX_PATH):
        print(f"[ERROR] ファイルが見つかりません: {XLSX_PATH}", file=sys.stderr)
        return 1
    try:
        wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    except Exception as e:
        print(f"[ERROR] Excel読込失敗: {e}", file=sys.stderr)
        return 2
    found = False
    for ws in wb.worksheets:
        header_row, word_idx, cefr_idx, pos_idx = detect_header_and_indices(ws)
        if header_row is None:
            continue
        found = True
        with open(OUT_PATH, "w", encoding="utf-8", newline="") as out:
            writer = csv.writer(out)
            writer.writerow(COLUMNS)
            for row in ws.iter_rows(min_row=header_row+1, values_only=True):
                try:
                    word = str(row[word_idx]).strip().lower() if row[word_idx] else ''
                    level = str(row[cefr_idx]).strip().upper() if row[cefr_idx] else ''
                    pos = str(row[pos_idx]).strip() if row[pos_idx] else ''
                    if not word or not level or not pos:
                        continue
                    writer.writerow([word, level, pos])
                except Exception:
                    continue
        print(f"Saved: {OUT_PATH}")
        break
    if not found:
        print("[ERROR] データ本体シートまたはヘッダーが検出できません", file=sys.stderr)
        return 3
    print("Done.")
    return 0

if __name__ == "__main__":
    try:
        import openpyxl
    except ImportError:
        import subprocess
        print("openpyxl not found, installing...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    sys.exit(main())
