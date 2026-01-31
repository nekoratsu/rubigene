import os
import sys
import csv
import zipfile
import io

ZIP_PATH = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "CEFRJ_wordlist_ver1.6.zip"))
OUT_PATH = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "data", "cefr.csv"))
COLUMNS = ["word", "CEFR_level", "pos"]

def ensure_openpyxl():
    try:
        import openpyxl
    except ImportError:
        import subprocess
        print("openpyxl not found, installing...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
        import openpyxl
    return openpyxl

def find_xlsx_file(zipf):
    for name in zipf.namelist():
        if name.lower().endswith('.xlsx'):
            return name
    return None


def extract_and_parse_xlsx(zipf, xlsx_name, out_path):
    openpyxl = ensure_openpyxl()
    try:
        with zipf.open(xlsx_name) as f:
            inmem = io.BytesIO(f.read())
            wb = openpyxl.load_workbook(inmem, read_only=True, data_only=True)
            found = False
            for ws in wb.worksheets:
                rows = ws.iter_rows(values_only=True)
                for i, row in enumerate(rows, 1):
                    if not row:
                        continue
                    headers = [str(cell).strip() if cell else '' for cell in row]
                    # 必須カラムが揃っているか
                    try:
                        col_idx = {k: headers.index(k) for k in ["word", "CEFR_level", "pos"]}
                    except ValueError:
                        continue
                    # データ本体発見
                    found = True
                    with open(out_path, "w", encoding="utf-8", newline="") as out:
                        writer = csv.writer(out)
                        writer.writerow(COLUMNS)
                        for data_row in ws.iter_rows(min_row=i+1, values_only=True):
                            try:
                                word = str(data_row[col_idx["word"]]).strip().lower() if data_row[col_idx["word"]] else ''
                                level = str(data_row[col_idx["CEFR_level"]]).strip().upper() if data_row[col_idx["CEFR_level"]] else ''
                                pos = str(data_row[col_idx["pos"]]).strip() if data_row[col_idx["pos"]] else ''
                                if not word or not level or not pos:
                                    continue
                                writer.writerow([word, level, pos])
                            except Exception:
                                continue
                    print(f"Saved: {out_path}")
                    return True
            if not found:
                print(f"[ERROR] 必須カラムが見つかりません（全シート・全行探索済）", file=sys.stderr)
                return False
    except Exception as e:
        print(f"[ERROR] Excelパース失敗: {e}", file=sys.stderr)
        return False

def main():
    if not os.path.exists(ZIP_PATH):
        print(f"[ERROR] ZIPファイルが見つかりません: {ZIP_PATH}", file=sys.stderr)
        return 1
    try:
        with zipfile.ZipFile(ZIP_PATH, 'r') as zipf:
            xlsx_name = find_xlsx_file(zipf)
            if not xlsx_name:
                print("[ERROR] ZIP内にExcelファイルが見つかりません", file=sys.stderr)
                return 2
            print(f"Found Excel file: {xlsx_name}")
            ok = extract_and_parse_xlsx(zipf, xlsx_name, OUT_PATH)
            if not ok:
                print("[ERROR] パースまたは保存に失敗", file=sys.stderr)
                return 3
    except Exception as e:
        print(f"[ERROR] ZIP展開失敗: {e}", file=sys.stderr)
        return 4
    print("Done.")
    return 0

if __name__ == "__main__":
    sys.exit(main())
